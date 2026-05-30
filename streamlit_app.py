import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import db
from datetime import datetime
import openpyxl
from io import BytesIO
import os


def _normalize_date_to_yyyy_mm_dd(raw_date) -> str:
    if isinstance(raw_date, datetime):
        return raw_date.strftime("%Y-%m-%d")
    if raw_date is None:
        return datetime.now().strftime("%Y-%m-%d")

    raw = str(raw_date).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return datetime.now().strftime("%Y-%m-%d")


def _import_patients_from_excel_bytes(excel_bytes: bytes) -> tuple[int, int]:
    wb = openpyxl.load_workbook(filename=BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    rows: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)

    if len(rows) == 0:
        raise ValueError("File Excel rỗng.")

    headers = [str(h).strip().upper() if h is not None else "" for h in rows[0]]
    is_kaggle = any(h == "PATIENT_VISIT_IDENTIFIER" for h in headers)

    imported_count = 0
    skipped_count = 0

    if is_kaggle:
        idx_id = -1
        idx_age_above65 = -1
        idx_age_percentil = -1
        idx_gender = -1
        idx_hr = -1
        idx_spo2 = -1
        idx_icu = -1

        for i, h in enumerate(headers):
            if h == "PATIENT_VISIT_IDENTIFIER":
                idx_id = i
            elif h == "AGE_ABOVE65":
                idx_age_above65 = i
            elif h == "AGE_PERCENTIL":
                idx_age_percentil = i
            elif h == "GENDER":
                idx_gender = i
            elif h == "HEART_RATE_MEAN":
                idx_hr = i
            elif h == "OXYGEN_SATURATION_MEAN":
                idx_spo2 = i
            elif h == "ICU":
                idx_icu = i

        if idx_id == -1:
            raise ValueError("Không tìm thấy cột PATIENT_VISIT_IDENTIFIER.")

        patients_dict: dict[int, dict] = {}

        for i in range(1, len(rows)):
            row = rows[i]
            try:
                if len(row) <= idx_id or row[idx_id] is None:
                    continue

                p_id = int(float(row[idx_id]))
                if p_id not in patients_dict:
                    patients_dict[p_id] = {
                        "age_above65": 0,
                        "age_percentil": "",
                        "gender": 0,
                        "hr_list": [],
                        "spo2_list": [],
                        "icu_list": [],
                    }

                    if idx_age_above65 != -1 and len(row) > idx_age_above65:
                        patients_dict[p_id]["age_above65"] = row[idx_age_above65]
                    if idx_age_percentil != -1 and len(row) > idx_age_percentil and row[idx_age_percentil] is not None:
                        patients_dict[p_id]["age_percentil"] = str(row[idx_age_percentil])
                    if idx_gender != -1 and len(row) > idx_gender:
                        patients_dict[p_id]["gender"] = row[idx_gender]

                if idx_hr != -1 and len(row) > idx_hr and row[idx_hr] is not None:
                    patients_dict[p_id]["hr_list"].append(float(row[idx_hr]))
                if idx_spo2 != -1 and len(row) > idx_spo2 and row[idx_spo2] is not None:
                    patients_dict[p_id]["spo2_list"].append(float(row[idx_spo2]))
                if idx_icu != -1 and len(row) > idx_icu and row[idx_icu] is not None:
                    patients_dict[p_id]["icu_list"].append(int(float(row[idx_icu])))
            except Exception:
                skipped_count += 1

        from datetime import timedelta

        for p_id, data in patients_dict.items():
            try:
                age = 35
                ap = (data.get("age_percentil") or "").strip()
                if ap:
                    ap_upper = ap.upper()
                    if "10TH" in ap_upper:
                        age = 15
                    elif "20TH" in ap_upper:
                        age = 25
                    elif "30TH" in ap_upper:
                        age = 35
                    elif "40TH" in ap_upper:
                        age = 45
                    elif "50TH" in ap_upper:
                        age = 55
                    elif "60TH" in ap_upper:
                        age = 65
                    elif "70TH" in ap_upper:
                        age = 75
                    elif "80TH" in ap_upper:
                        age = 85
                    elif "90TH" in ap_upper:
                        age = 90
                    elif "ABOVE 90" in ap_upper:
                        age = 95
                else:
                    if data.get("age_above65") == 1:
                        age = 70

                gender = "Nam" if data.get("gender") == 0 else "Nữ"

                icu = 1 if any(val == 1 for val in data.get("icu_list", [])) else 0

                hr_list = data.get("hr_list", [])
                if len(hr_list) > 0:
                    avg_hr_norm = sum(hr_list) / len(hr_list)
                    heart_rate = int((avg_hr_norm + 1) / 2 * (120 - 50) + 50)
                    heart_rate = max(30, min(250, heart_rate))
                else:
                    heart_rate = 80

                spo2_list = data.get("spo2_list", [])
                if len(spo2_list) > 0:
                    min_spo2_norm = min(spo2_list)
                    oxygen_saturation = int((min_spo2_norm + 1) / 2 * (100 - 80) + 80)
                    oxygen_saturation = max(50, min(100, oxygen_saturation))
                else:
                    oxygen_saturation = 98

                day_diff = p_id % 15
                admission_date = (datetime.now() - timedelta(days=day_diff)).strftime("%Y-%m-%d")
                name = f"Bệnh nhân Sirio #{p_id}"

                db.add_patient(name, age, gender, icu, heart_rate, oxygen_saturation, admission_date)
                imported_count += 1
            except Exception:
                skipped_count += 1

        return imported_count, skipped_count

    for i in range(1, len(rows)):
        row = rows[i]

        is_empty = True
        for cell in row:
            if cell is not None:
                is_empty = False
                break
        if is_empty:
            continue

        try:
            name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else "Bệnh nhân ẩn danh"
            age = int(row[1]) if len(row) > 1 and row[1] is not None else 0
            gender = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "Không rõ"

            icu = int(row[3]) if len(row) > 3 and row[3] is not None else 0
            icu = 1 if icu > 0 else 0

            heart_rate = int(row[4]) if len(row) > 4 and row[4] is not None else 80
            oxygen_saturation = int(row[5]) if len(row) > 5 and row[5] is not None else 98

            raw_date = row[6] if len(row) > 6 else None
            admission_date = _normalize_date_to_yyyy_mm_dd(raw_date)

            db.add_patient(name, age, gender, icu, heart_rate, oxygen_saturation, admission_date)
            imported_count += 1
        except Exception:
            skipped_count += 1

    return imported_count, skipped_count

# Streamlit Page Config
st.set_page_config(page_title="MedICU - Dashboard", page_icon="🏥", layout="wide")

# Ensure DB exists when running via Streamlit
db.init_db()


@st.cache_data
def _build_sample_excel_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách bệnh nhân"

    headers = [
        "Họ và Tên",
        "Tuổi",
        "Giới tính",
        "ICU (1=Có, 0=Không)",
        "Nhịp tim",
        "SpO2 (%)",
        "Ngày nhập viện (YYYY-MM-DD)",
    ]
    ws.append(headers)
    ws.append(["Nguyễn Văn Anh", 45, "Nam", 0, 75, 98, "2026-05-20"])
    ws.append(["Trần Thị Mai", 72, "Nữ", 1, 95, 91, "2026-05-20"])
    ws.append(["Lê Hoàng Nam", 19, "Nam", 0, 80, 99, "2026-05-20"])

    file_stream = BytesIO()
    wb.save(file_stream)
    return file_stream.getvalue()

# Sidebar navigation
st.sidebar.title("🏥 MedICU")
st.sidebar.subheader("Hệ Thống Phân Tích Bệnh Nhân")

menu = [
    "Tổng Quan",
    "Bản Đồ & Ma Trận Rủi Ro",
    "Chuỗi Thời Gian & Dự Báo",
    "Phân Tích Sâu",
    "Dữ Liệu Chi Tiết",
    "Hướng Dẫn"
]
choice = st.sidebar.radio("Điều hướng", menu)

st.sidebar.markdown("---")
st.sidebar.markdown("**Tải & Nhập Excel**")

st.sidebar.download_button(
    label="Tải Excel mẫu",
    data=_build_sample_excel_bytes(),
    file_name="mau_du_lieu_benh_nhan.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    width="stretch",
)

_kaggle_path = os.path.join(os.path.dirname(__file__), "Kaggle_Sirio_Libanes_ICU_Prediction.xlsx")
if os.path.exists(_kaggle_path):
    with open(_kaggle_path, "rb") as f:
        st.sidebar.download_button(
            label="Tải Excel Kaggle (Sirio Libanes)",
            data=f.read(),
            file_name=os.path.basename(_kaggle_path),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

uploaded_file = st.sidebar.file_uploader("Upload Excel", type=["xlsx", "xls"])
if uploaded_file is not None:
    st.sidebar.caption(f"Đã chọn: {uploaded_file.name}")
    if st.sidebar.button("Import Dữ Liệu", type="primary", width="stretch"):
        try:
            imported, skipped = _import_patients_from_excel_bytes(uploaded_file.getvalue())
            st.sidebar.success(f"Nhập thành công {imported} bệnh nhân. Bỏ qua {skipped} dòng lỗi.")
            st.rerun()
        except Exception as e:
            st.sidebar.error(f"Lỗi xử lý file: {e}")

st.sidebar.markdown("---")
st.sidebar.success("SQLite Connected")

if choice == "Tổng Quan":
    st.title("Dashboard Phân Tích Bệnh Nhân ICU")
    st.markdown("Dữ liệu phân tích lâm sàng và tỷ lệ tiếp nhận phòng điều trị tích cực")
    
    stats = db.get_dashboard_stats()
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Tổng số bệnh nhân", stats["total_patients"])
    col2.metric("Nhập khoa ICU", stats["icu_count"])
    col3.metric("Tỷ lệ ICU", f"{stats['icu_rate']}%")
    col4.metric("Nhịp tim / SpO2 TB", f"{stats['avg_heart_rate']} / {stats['avg_oxygen']}%")
    
    st.markdown("---")
    
    row1_col1, row1_col2 = st.columns(2)
    
    with row1_col1:
        st.subheader("Phân Phối Ca Bệnh Điều Trị")
        pie_data = stats["pie_data"]
        fig_pie = px.pie(
            names=["ICU", "Không ICU"], 
            values=[pie_data["icu"], pie_data["non_icu"]],
            color_discrete_sequence=["#f43f5e", "#6366f1"]
        )
        fig_pie.update_traces(hole=.6)
        st.plotly_chart(fig_pie, width="stretch")
        
    with row1_col2:
        st.subheader("Tỷ Lệ Nhập ICU Theo Nhóm Tuổi")
        bar_data = pd.DataFrame(stats["bar_data"])
        if not bar_data.empty:
            fig_bar = go.Figure()
            fig_bar.add_trace(go.Bar(x=bar_data["age_group"] + " tuổi", y=bar_data["total_patients"], name="Tổng", marker_color="#6366f1"))
            fig_bar.add_trace(go.Bar(x=bar_data["age_group"] + " tuổi", y=bar_data["icu_patients"], name="ICU", marker_color="#f43f5e"))
            
            fig_bar.add_trace(go.Scatter(
                x=bar_data["age_group"] + " tuổi", 
                y=bar_data["icu_rate"], 
                name="Tỷ lệ ICU (%)", 
                yaxis="y2", 
                line=dict(color="#f97316", width=3)
            ))
            
            fig_bar.update_layout(
                yaxis=dict(title="Số bệnh nhân"),
                yaxis2=dict(title="Tỷ lệ ICU (%)", overlaying="y", side="right", range=[0, 100]),
                barmode='group'
            )
            st.plotly_chart(fig_bar, width="stretch")
            
    st.subheader("Chỉ Số Sức Khỏe Trung Bình Theo Độ Tuổi")
    line_data = pd.DataFrame(stats["line_data"])
    if not line_data.empty:
        fig_line = go.Figure()
        fig_line.add_trace(go.Scatter(
            x=line_data["age_group"] + " tuổi", 
            y=line_data["avg_heart_rate"],
            name="Nhịp tim (BPM)",
            line=dict(color="#f43f5e", width=3),
            yaxis="y1"
        ))
        fig_line.add_trace(go.Scatter(
            x=line_data["age_group"] + " tuổi", 
            y=line_data["avg_oxygen_saturation"],
            name="SpO2 (%)",
            line=dict(color="#14b8a6", width=3),
            yaxis="y2"
        ))
        fig_line.update_layout(
            yaxis=dict(title="Nhịp tim (BPM)"),
            yaxis2=dict(title="SpO2 (%)", overlaying="y", side="right")
        )
        st.plotly_chart(fig_line, width="stretch")


elif choice == "Bản Đồ & Ma Trận Rủi Ro":
    st.title("Bản Đồ & Ma Trận Rủi Ro")
    st.markdown("Phân loại rủi ro theo SpO2, nhịp tim và tỷ lệ ICU")
    
    risk_data = db.get_risk_matrix_data()
    df_risk = pd.DataFrame(risk_data)
    
    if not df_risk.empty:
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.subheader("Bản đồ Rủi ro (Treemap)")
            # Calculate size (count) and color (ICU rate) for treemap
            grouped = df_risk.groupby(['spo2_cat', 'hr_cat']).agg(
                count=('id', 'count'),
                icu_rate=('icu', 'mean')
            ).reset_index()
            
            # Map category names to Vietnamese for better UI
            hr_map = {"bradycardia": "Chậm (<60)", "normal": "Bình thường (60-100)", "tachycardia": "Nhanh (>100)"}
            spo2_map = {"high": "Cao (≥95%)", "medium": "Trung bình (90-94%)", "low": "Thấp (<90%)"}
            
            grouped['Nhịp tim'] = grouped['hr_cat'].map(hr_map)
            grouped['SpO2'] = grouped['spo2_cat'].map(spo2_map)
            
            fig_tree = px.treemap(
                grouped, 
                path=['SpO2', 'Nhịp tim'], 
                values='count',
                color='icu_rate',
                color_continuous_scale='Reds',
                labels={'icu_rate': 'Tỷ lệ ICU', 'count': 'Số ca'}
            )
            fig_tree.update_traces(textinfo="label+value")
            fig_tree.update_layout(height=450, margin=dict(t=30, l=10, r=10, b=10))
            st.plotly_chart(fig_tree, width="stretch")
            
        with col2:
            st.subheader("Ma Trận Rủi Ro (Heatmap)")
            # Create a 3x3 matrix for heatmap
            matrix_data = pd.DataFrame(0, 
                index=["Cao (≥95%)", "Trung bình (90-94%)", "Thấp (<90%)"], 
                columns=["Chậm (<60)", "Bình thường (60-100)", "Nhanh (>100)"]
            )
            
            for index, row in grouped.iterrows():
                matrix_data.at[row['SpO2'], row['Nhịp tim']] = row['count']
                
            fig_heat = px.imshow(
                matrix_data,
                labels=dict(x="Nhịp tim", y="SpO2", color="Số lượng BN"),
                x=matrix_data.columns,
                y=matrix_data.index,
                text_auto=True,
                color_continuous_scale="Greens" if matrix_data.values.max() > 0 else "gray"
            )
            fig_heat.update_xaxes(side="top")
            fig_heat.update_layout(height=450, margin=dict(t=30, l=10, r=10, b=10))
            st.plotly_chart(fig_heat, width="stretch")
            
        st.markdown("---")
        st.subheader("Dữ liệu chi tiết rủi ro")
        df_display = df_risk[['name', 'age', 'heart_rate', 'oxygen_saturation', 'spo2_cat', 'hr_cat', 'icu']].copy()
        df_display['hr_cat'] = df_display['hr_cat'].map(hr_map)
        df_display['spo2_cat'] = df_display['spo2_cat'].map(spo2_map)
        df_display.columns = ["Họ và tên", "Tuổi", "Nhịp tim", "SpO2", "Nhóm SpO2", "Nhóm Nhịp tim", "ICU"]
        st.dataframe(df_display, width="stretch")
    else:
        st.info("Chưa có dữ liệu bệnh nhân.")

elif choice == "Chuỗi Thời Gian & Dự Báo":
    st.title("Chuỗi Thời Gian & Dự Báo")
    
    ts_data = db.get_time_series_data()
    df_ts = pd.DataFrame(ts_data)
    
    if not df_ts.empty:
        st.subheader("Diễn biến Chuỗi thời gian")
        fig_ts = px.line(df_ts, x="admission_date", y=["patient_count", "avg_heart_rate", "avg_oxygen_saturation"],
                         labels={"value": "Giá trị", "admission_date": "Ngày nhập viện"},
                         title="Số lượng nhập viện & Sinh tồn TB")
        st.plotly_chart(fig_ts, width="stretch")
        
        st.subheader("Dự báo Số lượng Nhập viện (Hồi quy)")
        fc_data = db.get_forecast_data()
        df_fc = pd.DataFrame(fc_data)
        if not df_fc.empty:
            fig_fc = px.line(df_fc, x="admission_date", y="patient_count", title="Dự báo 7 ngày tiếp theo", line_shape="spline")
            fig_fc.update_traces(line_color="#a855f7")
            st.plotly_chart(fig_fc, width="stretch")
    else:
        st.info("Chưa có dữ liệu.")

elif choice == "Phân Tích Sâu":
    st.title("Phân Tích Sâu (Nâng Cao)")
    adv_stats = db.get_advanced_stats()

    def _interpret_corr(r: float) -> str:
        abs_r = abs(r)
        direction = "cùng chiều" if r > 0 else ("ngược chiều" if r < 0 else "trung tính")
        if abs_r < 0.10:
            strength = "rất yếu"
        elif abs_r < 0.30:
            strength = "yếu"
        elif abs_r < 0.50:
            strength = "vừa"
        else:
            strength = "mạnh"
        return f"{strength} ({direction})"

    def _format_r(r: float) -> str:
        return f"{r:.3f}" if isinstance(r, (int, float)) else str(r)
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Tương quan Tuổi vs SpO2 (r)", _format_r(adv_stats["age_spo2_corr"]))
    col2.metric("Tương quan Tuổi vs Nhịp tim (r)", _format_r(adv_stats["age_hr_corr"]))
    col3.metric("Rủi ro ICU người lớn tuổi", f"{adv_stats['relative_risk']}x")

    st.caption(
        "Gợi ý đọc nhanh: r gần 0 → mối liên hệ tuyến tính yếu; |r| càng lớn → mối liên hệ tuyến tính càng rõ. "
        "Dấu âm/dương cho biết xu hướng ngược/cùng chiều."
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"- Tuổi vs SpO2: **{_interpret_corr(float(adv_stats['age_spo2_corr']))}**")
    with c2:
        st.markdown(f"- Tuổi vs Nhịp tim: **{_interpret_corr(float(adv_stats['age_hr_corr']))}**")

    with st.expander("Giải thích cách tính (phương pháp)", expanded=False):
        st.markdown("**1) Dữ liệu đầu vào**")
        st.markdown(
            "- Lấy toàn bộ bệnh nhân trong SQLite (bảng `patients`).\n"
            "- Biến dùng cho tương quan: Tuổi (age), SpO2 (oxygen_saturation), Nhịp tim (heart_rate)."
        )

        st.markdown("**2) Tương quan Pearson (r)**")
        st.markdown("Tính hệ số tương quan Pearson để xem mức độ liên hệ tuyến tính giữa 2 biến số.")
        st.latex(
            r"r=\frac{\sum_{i=1}^{n}(x_i-\bar{x})(y_i-\bar{y})}{\sqrt{\sum_{i=1}^{n}(x_i-\bar{x})^2}\sqrt{\sum_{i=1}^{n}(y_i-\bar{y})^2}}"
        )
        st.markdown(
            "- $r \\in [-1, 1]$: gần 1 (cùng chiều mạnh), gần -1 (ngược chiều mạnh), gần 0 (liên hệ tuyến tính yếu).\n"
            "- Đây là mô tả thống kê; **không** khẳng định quan hệ nhân-quả."
        )

        st.markdown("**3) Rủi ro ICU người lớn tuổi (Relative Risk)**")
        st.markdown("So sánh tỷ lệ ICU giữa 2 nhóm: tuổi ≥ 60 và < 60.")
        st.latex(
            r"RR=\frac{P(ICU\mid age\ge 60)}{P(ICU\mid age<60)}"
        )
        st.markdown(
            "- RR > 1: nhóm ≥60 có tỷ lệ ICU cao hơn.\n"
            "- RR = 1: tương đương.\n"
            "- RR < 1: nhóm ≥60 thấp hơn."
        )

        elderly_total = int(adv_stats.get("elderly_total", 0))
        younger_total = int(adv_stats.get("younger_total", 0))
        st.markdown("**4) Quy mô mẫu (để bạn đối chiếu độ tin cậy)**")
        st.markdown(
            f"- Số BN tuổi ≥60: **{elderly_total}**\n"
            f"- Số BN tuổi <60: **{younger_total}**"
        )
        if elderly_total + younger_total < 30:
            st.info("Cỡ mẫu còn nhỏ; các chỉ số có thể dao động nhiều khi thêm dữ liệu.")
    
    st.markdown("### Báo cáo Phân tích Phân nhóm Lâm sàng & Dịch tễ học")
    st.write(f"- Tỷ lệ nhập ICU ở người trên 60 tuổi: **{adv_stats['elderly_icu_risk']}%**")
    st.write(f"- Tỷ lệ nhập ICU ở người dưới 60 tuổi: **{adv_stats['younger_icu_risk']}%**")

elif choice == "Dữ Liệu Chi Tiết":
    st.title("Danh Sách Bệnh Nhân ICU")

    col_a, col_b = st.columns([1, 3])
    with col_a:
        if st.button("🗑️ Xóa tất cả", type="secondary", width="stretch"):
            db.delete_all_patients()
            st.success("Đã xóa toàn bộ bệnh nhân.")
            st.rerun()
    
    with st.expander("➕ Thêm Bệnh Nhân Mới"):
        with st.form("add_patient_form"):
            name = st.text_input("Họ và Tên")
            age = st.number_input("Tuổi", min_value=0, max_value=120, value=30)
            gender = st.selectbox("Giới tính", ["Nam", "Nữ", "Khác"])
            heart_rate = st.number_input("Nhịp tim (BPM)", min_value=30, max_value=250, value=80)
            oxygen = st.number_input("SpO2 (%)", min_value=50, max_value=100, value=98)
            admission_date = st.date_input("Ngày nhập viện", datetime.today())
            icu = st.radio("Điều trị tại khoa ICU?", ["Có", "Không"])
            
            submitted = st.form_submit_button("Lưu thông tin")
            if submitted:
                icu_val = 1 if icu == "Có" else 0
                db.add_patient(name, age, gender, icu_val, heart_rate, oxygen, admission_date.strftime("%Y-%m-%d"))
                st.success("Thêm bệnh nhân thành công!")
                
    st.subheader("Dữ Liệu")
    patients = db.get_all_patients()
    df_patients = pd.DataFrame(patients)
    
    if not df_patients.empty:
        st.dataframe(df_patients, width="stretch")
        
        st.markdown("### Xóa bệnh nhân")
        del_id = st.number_input("Nhập ID bệnh nhân cần xóa", min_value=0, step=1)
        if st.button("Xóa"):
            if db.delete_patient(del_id):
                st.success(f"Đã xóa bệnh nhân ID {del_id}")
                st.rerun()
            else:
                st.error("Không tìm thấy ID")
    else:
        st.info("Chưa có bệnh nhân nào.")

elif choice == "Hướng Dẫn":
    st.title("Hướng Dẫn Sử Dụng")
    st.markdown("""
    ### Quản Lý Dữ Liệu Bệnh Nhân
    Ứng dụng hỗ trợ lưu trữ dữ liệu bệnh nhân thông qua hệ quản trị cơ sở dữ liệu SQLite.
    
    - **Thêm mới:** Vào mục 'Dữ Liệu Chi Tiết', mở rộng phần Thêm Bệnh Nhân Mới và điền thông tin.
    - **Xóa:** Dùng chức năng xóa phía dưới danh sách bệnh nhân.
    
    ### Dashboard & Báo cáo
    - Tự động thống kê và vẽ biểu đồ.
    - Tích hợp mô hình hồi quy tuyến tính dự báo số lượng.
    - Đánh giá chỉ số rủi ro dựa vào SpO2 và Nhịp tim.
    """)
