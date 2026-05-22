from fastapi import FastAPI, HTTPException, UploadFile, File, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
import os
import openpyxl
from io import BytesIO
from datetime import datetime

# Import file db.py do chúng ta tự viết
import db

# Khởi tạo ứng dụng FastAPI
app = FastAPI(title="ICU Patient Dashboard API")

# Cấu hình để cho phép frontend gọi được API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Chạy lệnh này khi bắt đầu bật server
@app.on_event("startup")
def startup_event():
    db.init_db()

# Cấu trúc dữ liệu nhận từ người dùng
class PatientCreate(BaseModel):
    name: str
    age: int
    gender: str
    icu: int
    heart_rate: int
    oxygen_saturation: int
    admission_date: str = ""

@app.get("/ping")
def ping():
    return {"status": "ok", "message": "Backend is running!"}

@app.get("/api/patients")
def read_patients():
    try:
        patients = db.get_all_patients()
        return patients
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.post("/api/patients")
def create_patient(patient: PatientCreate):
    try:
        # Nếu không có ngày nhập viện, dùng ngày hôm nay
        if patient.admission_date == "":
            now = datetime.now()
            patient.admission_date = now.strftime("%Y-%m-%d")
        else:
            # Thử xem ngày nhập viện có đúng định dạng không
            try:
                datetime.strptime(patient.admission_date, "%Y-%m-%d")
            except Exception:
                # Nếu sai, lại dùng ngày hôm nay
                now = datetime.now()
                patient.admission_date = now.strftime("%Y-%m-%d")

        patient_id = db.add_patient(
            patient.name,
            patient.age,
            patient.gender,
            patient.icu,
            patient.heart_rate,
            patient.oxygen_saturation,
            patient.admission_date
        )
        return {"status": "success", "id": patient_id, "message": "Đã thêm bệnh nhân thành công."}
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.delete("/api/patients/{patient_id}")
def delete_patient(patient_id: int):
    try:
        success = db.delete_patient(patient_id)
        if success == False:
            raise HTTPException(status_code=404, detail="Không tìm thấy bệnh nhân để xóa.")
        return {"status": "success", "message": "Đã xóa bệnh nhân thành công."}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.delete("/api/patients")
def delete_all_patients():
    try:
        db.delete_all_patients()
        return {"status": "success", "message": "Đã xóa toàn bộ bệnh nhân thành công."}
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.get("/api/patients/stats")
def read_stats():
    try:
        stats = db.get_dashboard_stats()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.get("/api/patients/time-series")
def read_time_series():
    try:
        data = db.get_time_series_data()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.get("/api/patients/risk-matrix")
def read_risk_matrix():
    try:
        data = db.get_risk_matrix_data()
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.get("/api/patients/advanced-stats")
def read_advanced_stats():
    try:
        stats = db.get_advanced_stats()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.get("/api/patients/forecast")
def read_forecast():
    try:
        forecast = db.get_forecast_data()
        return forecast
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.get("/api/patients/sample-excel")
def download_sample_excel():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách bệnh nhân"
        
        headers = ["Họ và Tên", "Tuổi", "Giới tính", "ICU (1=Có, 0=Không)", "Nhịp tim", "SpO2 (%)", "Ngày nhập viện (YYYY-MM-DD)"]
        ws.append(headers)
        
        ws.append(["Nguyễn Văn Anh", 45, "Nam", 0, 75, 98, "2026-05-20"])
        ws.append(["Trần Thị Mai", 72, "Nữ", 1, 95, 91, "2026-05-20"])
        ws.append(["Lê Hoàng Nam", 19, "Nam", 0, 80, 99, "2026-05-20"])
        
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=mau_du_lieu_benh_nhan.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi: " + str(e))

@app.get("/api/patients/kaggle-excel")
def download_kaggle_excel():
    file_path = os.path.join(os.path.dirname(__file__), "Kaggle_Sirio_Libanes_ICU_Prediction.xlsx")
    if os.path.exists(file_path) == False:
        raise HTTPException(status_code=404, detail="Không tìm thấy tệp.")
        
    def iterfile():
        f = open(file_path, mode="rb")
        yield from f
        f.close()
            
    filename = os.path.basename(file_path)
    return StreamingResponse(
        iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + filename}
    )

@app.post("/api/patients/import")
async def import_patients_excel(file: UploadFile = File(...)):
    if file.filename.endswith(".xlsx") == False and file.filename.endswith(".xls") == False:
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ file Excel")
        
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
        ws = wb.active
        
        imported_count = 0
        skipped_count = 0
        
        is_kaggle = False
        
        # Lấy tất cả các dòng vào một mảng list
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(row)
            
        if len(rows) == 0:
            raise HTTPException(status_code=400, detail="File Excel rỗng.")
            
        headers = []
        for h in rows[0]:
            if h is not None:
                headers.append(str(h).strip().upper())
            else:
                headers.append("")
                
        # Kiểm tra xem có phải file Kaggle không
        for h in headers:
            if h == "PATIENT_VISIT_IDENTIFIER":
                is_kaggle = True
                break
                
        if is_kaggle == True:
            try:
                idx_id = -1
                idx_age_above65 = -1
                idx_age_percentil = -1
                idx_gender = -1
                idx_hr = -1
                idx_spo2 = -1
                idx_icu = -1
                
                for i in range(len(headers)):
                    if headers[i] == "PATIENT_VISIT_IDENTIFIER": idx_id = i
                    elif headers[i] == "AGE_ABOVE65": idx_age_above65 = i
                    elif headers[i] == "AGE_PERCENTIL": idx_age_percentil = i
                    elif headers[i] == "GENDER": idx_gender = i
                    elif headers[i] == "HEART_RATE_MEAN": idx_hr = i
                    elif headers[i] == "OXYGEN_SATURATION_MEAN": idx_spo2 = i
                    elif headers[i] == "ICU": idx_icu = i
                    
                patients_dict = {}
                
                # Duyệt từ dòng 2 trở đi
                for i in range(1, len(rows)):
                    row = rows[i]
                    if len(row) <= idx_id or row[idx_id] is None:
                        continue
                        
                    p_id = int(row[idx_id])
                    
                    if p_id not in patients_dict:
                        patients_dict[p_id] = {
                            "age_above65": 0,
                            "age_percentil": "",
                            "gender": 0,
                            "hr_list": [],
                            "spo2_list": [],
                            "icu_list": []
                        }
                        if idx_age_above65 != -1 and len(row) > idx_age_above65:
                            patients_dict[p_id]["age_above65"] = row[idx_age_above65]
                        if idx_age_percentil != -1 and len(row) > idx_age_percentil and row[idx_age_percentil] is not None:
                            patients_dict[p_id]["age_percentil"] = str(row[idx_age_percentil])
                        if idx_gender != -1 and len(row) > idx_gender:
                            patients_dict[p_id]["gender"] = row[idx_gender]
                            
                    if idx_hr != -1 and len(row) > idx_hr and row[idx_hr] is not None:
                        patients_dict[p_id]["hr_list"].append(float(row[idx_hr]))
                    if idx_spo2 != -1 and len(row) > idx_spo2 and row[idx_spo2] is not None:
                        patients_dict[p_id]["spo2_list"].append(float(row[idx_spo2]))
                    if idx_icu != -1 and len(row) > idx_icu and row[idx_icu] is not None:
                        patients_dict[p_id]["icu_list"].append(int(row[idx_icu]))
                        
                for p_id in patients_dict:
                    data = patients_dict[p_id]
                    try:
                        age = 35
                        ap = data["age_percentil"]
                        if ap != "":
                            ap_upper = ap.upper()
                            if "10TH" in ap_upper: age = 15
                            elif "20TH" in ap_upper: age = 25
                            elif "30TH" in ap_upper: age = 35
                            elif "40TH" in ap_upper: age = 45
                            elif "50TH" in ap_upper: age = 55
                            elif "60TH" in ap_upper: age = 65
                            elif "70TH" in ap_upper: age = 75
                            elif "80TH" in ap_upper: age = 85
                            elif "90TH" in ap_upper: age = 90
                            elif "ABOVE 90" in ap_upper: age = 95
                        else:
                            if data["age_above65"] == 1:
                                age = 70
                                
                        if data["gender"] == 0:
                            gender = "Nam"
                        else:
                            gender = "Nữ"
                            
                        icu = 0
                        for val in data["icu_list"]:
                            if val == 1:
                                icu = 1
                                
                        if len(data["hr_list"]) > 0:
                            sum_hr = 0
                            for hr in data["hr_list"]:
                                sum_hr = sum_hr + hr
                            avg_hr_norm = sum_hr / len(data["hr_list"])
                            heart_rate = int((avg_hr_norm + 1) / 2 * (120 - 50) + 50)
                            if heart_rate < 30: heart_rate = 30
                            if heart_rate > 250: heart_rate = 250
                        else:
                            heart_rate = 80
                            
                        if len(data["spo2_list"]) > 0:
                            min_spo2_norm = data["spo2_list"][0]
                            for spo2 in data["spo2_list"]:
                                if spo2 < min_spo2_norm:
                                    min_spo2_norm = spo2
                            oxygen_saturation = int((min_spo2_norm + 1) / 2 * (100 - 80) + 80)
                            if oxygen_saturation < 50: oxygen_saturation = 50
                            if oxygen_saturation > 100: oxygen_saturation = 100
                        else:
                            oxygen_saturation = 98
                            
                        from datetime import timedelta
                        day_diff = p_id % 15
                        now = datetime.now()
                        admission_date = (now - timedelta(days=day_diff)).strftime("%Y-%m-%d")
                        name = "Bệnh nhân Sirio #" + str(p_id)
                        
                        db.add_patient(name, age, gender, icu, heart_rate, oxygen_saturation, admission_date)
                        imported_count = imported_count + 1
                    except Exception as e:
                        skipped_count = skipped_count + 1
            except Exception as e:
                raise HTTPException(status_code=400, detail="Lỗi xử lý file Kaggle: " + str(e))
        else:
            for i in range(1, len(rows)):
                row = rows[i]
                
                is_empty = True
                for cell in row:
                    if cell is not None:
                        is_empty = False
                if is_empty == True:
                    continue
                    
                try:
                    if row[0] is not None: name = str(row[0]).strip()
                    else: name = "Bệnh nhân ẩn danh"
                    
                    if row[1] is not None: age = int(row[1])
                    else: age = 0
                    
                    if row[2] is not None: gender = str(row[2]).strip()
                    else: gender = "Không rõ"
                    
                    if row[3] is not None: icu = int(row[3])
                    else: icu = 0
                    
                    if icu > 0: icu = 1
                    else: icu = 0
                    
                    if row[4] is not None: heart_rate = int(row[4])
                    else: heart_rate = 80
                    
                    if row[5] is not None: oxygen_saturation = int(row[5])
                    else: oxygen_saturation = 98
                    
                    raw_date = row[6]
                    if isinstance(raw_date, datetime):
                        admission_date = raw_date.strftime("%Y-%m-%d")
                    elif raw_date is not None:
                        raw_date_str = str(raw_date).strip()
                        try:
                            admission_date = datetime.strptime(raw_date_str, "%Y-%m-%d").strftime("%Y-%m-%d")
                        except Exception:
                            try:
                                admission_date = datetime.strptime(raw_date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
                            except Exception:
                                admission_date = datetime.now().strftime("%Y-%m-%d")
                    else:
                        admission_date = datetime.now().strftime("%Y-%m-%d")
                        
                    db.add_patient(name, age, gender, icu, heart_rate, oxygen_saturation, admission_date)
                    imported_count = imported_count + 1
                except Exception as e:
                    skipped_count = skipped_count + 1
                    
        return {
            "status": "success", 
            "message": "Nhập thành công " + str(imported_count) + " bệnh nhân. Bỏ qua " + str(skipped_count) + " dòng lỗi.",
            "imported": imported_count,
            "skipped": skipped_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail="Lỗi xử lý file: " + str(e))

frontend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "python-dashboard-fe"))
if os.path.exists(frontend_dir):
    static_dir = os.path.join(frontend_dir, "static")
    if os.path.exists(static_dir):
        app.mount("/static", StaticFiles(directory=static_dir), name="static")
        
    templates_dir = os.path.join(frontend_dir, "templates")
    if os.path.exists(templates_dir):
        templates = Jinja2Templates(directory=templates_dir)
        
        @app.get("/")
        async def home(request: Request, tab: str = "dashboard"):
            stats = db.get_dashboard_stats()
            patients = db.get_all_patients()
            return templates.TemplateResponse(
                request=request,
                name="index.html",
                context={
                    "tab": tab,
                    "stats": stats,
                    "patients": patients
                }
            )
            
        @app.post("/create")
        async def create_patient_form(
            name: str = Form(...),
            age: int = Form(...),
            gender: str = Form(...),
            icu: int = Form(...),
            heart_rate: int = Form(...),
            oxygen_saturation: int = Form(...),
            admission_date: str = Form(...)
        ):
            db.add_patient(name, age, gender, icu, heart_rate, oxygen_saturation, admission_date)
            return RedirectResponse(url="/?tab=patients", status_code=303)
            
        @app.get("/delete/{patient_id}")
        async def delete_patient_route(patient_id: int):
            db.delete_patient(patient_id)
            return RedirectResponse(url="/?tab=patients", status_code=303)
else:
    print("Warning: Frontend directory not found at " + frontend_dir)
